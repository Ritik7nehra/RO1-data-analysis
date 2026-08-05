"""Data loading and cleaning utilities for the clinical visit dashboard.

The source Excel workbook is never modified. The dashboard builds an in-memory,
longitudinal table with one row per patient/visit and a normalized ``Visit`` label.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook


MISSING_TOKENS = {
    "",
    "na",
    "n/a",
    "nan",
    "none",
    "null",
    "missing",
    "not available",
    "not applicable",
    "unknown",
}

BASE_SHEET_HINTS = (
    "base",
    "master",
    "raw",
    "all data",
    "alldata",
    "full data",
    "source",
)

META_COLUMNS = {
    "Patient ID",
    "Event Name",
    "Visit",
    "Visit Date",
    "Source Sheet",
    "Source Row",
    "Duplicate Patient-Visit",
}

# Edit or extend these patterns when a new workbook uses different labels.
# The dashboard also exposes detected columns in the UI so metrics can be selected
# without changing code.
CANONICAL_COLUMN_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"^(record|study|patient|subject)\s*(id|identifier)$", re.I), "Patient ID"),
    (re.compile(r"^(event|visit)\s*(name|label|stage)?$", re.I), "Event Name"),
    (re.compile(r"^date\s*of\s*birth\d*$|^dob\d*$", re.I), "Date of Birth"),
    (re.compile(r"^(age|age\s*at\s*visit)$", re.I), "Age"),
    (re.compile(r"^calculated\s*bmi$|^bmi$", re.I), "Calculated BMI"),
    (re.compile(r"^mean\s*sbp$|^systolic\s*blood\s*pressure$", re.I), "Mean SBP"),
    (re.compile(r"^mean\s*dbp$|^diastolic\s*blood\s*pressure$", re.I), "Mean DBP"),
    (re.compile(r"^mean\s*hr$|^mean\s*heart\s*rate$", re.I), "Mean HR"),
    (re.compile(r"^waist\s*hip\s*ratio$", re.I), "Waist Hip Ratio"),
    (re.compile(r"^use\s*of\s*cgm$", re.I), "Use of CGM"),
    (re.compile(r"^pump$", re.I), "Pump"),
    (re.compile(r"^diabetic\s*retinopathy$", re.I), "Diabetic Retinopathy"),
    (re.compile(r"^diabetic\s*neuropathy$", re.I), "Diabetic Neuropathy"),
    (re.compile(r"^diabetic\s*nephropathy$", re.I), "Diabetic Nephropathy"),
]

DEFAULT_METRIC_PRIORITY = [
    "HbA1c",
    "Calculated BMI",
    "Mean Weight (kg)",
    "Mean SBP",
    "Mean DBP",
    "Finger Stick Blood Glucose (mg/dL)",
    "Age",
    "Mean Waist (cm)",
    "Mean Hip (cm)",
    "Waist Hip Ratio",
    "Average HbA1c",
    "Readings per year before Visit1",
]


@dataclass(frozen=True)
class SheetInfo:
    name: str
    headers: tuple[Any, ...]
    max_row: int
    max_column: int


def _plain_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def _header_signature(value: Any) -> str:
    """Return a comparison-friendly header signature."""
    text = _plain_text(value)
    text = re.sub(r"\.\d+$", "", text)  # pandas duplicate-header suffix
    text = text.replace("&", " and ")
    text = text.replace("HbA1C", "HbA1c").replace("HBA1C", "HbA1c")
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"[():,;?]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def canonical_column_name(value: Any) -> str:
    """Normalize inconsistent headers while retaining meaningful labels."""
    raw = _plain_text(value)
    if not raw or raw.lower().startswith("unnamed:"):
        return ""

    signature = _header_signature(raw)

    for pattern, canonical in CANONICAL_COLUMN_PATTERNS:
        if pattern.fullmatch(signature):
            return canonical

    # Visit dates. Exclude other clinical dates such as DOB or diagnosis date.
    if "date" in signature and "visit" in signature:
        if not any(term in signature for term in ("birth", "diagnos", "collection", "cgm", "start", "stop")):
            return "Visit Date"

    # HbA1c aliases. Keep historical average columns separate from visit HbA1c.
    if ("hba1c" in signature or "a1c" in signature) and "control" not in signature:
        if "avg" in signature or "average" in signature:
            return "Average HbA1c"
        if "reading" in signature and "year" in signature:
            return "Readings per year before Visit1"
        if "value" in signature or signature in {"hba1c", "a1c"}:
            return "HbA1c"

    if "reading" in signature and "year" in signature and "visit1" in signature.replace(" ", ""):
        return "Readings per year before Visit1"

    if "finger" in signature and "glucose" in signature:
        return "Finger Stick Blood Glucose (mg/dL)"

    if "height" in signature and "loss" not in signature and "mean" in signature:
        return "Mean Height (cm)"
    if "weight" in signature and "pregnan" not in signature and "body weight" not in signature:
        if "mean" in signature or signature in {"weight kg", "weight"}:
            return "Mean Weight (kg)"
    if "waist" in signature and "ratio" not in signature and "mean" in signature:
        return "Mean Waist (cm)"
    if "hip" in signature and "ratio" not in signature and "mean" in signature:
        return "Mean Hip (cm)"

    # Preserve readable original labels after trimming spaces and trailing punctuation.
    cleaned = re.sub(r"\s+", " ", raw).strip()
    cleaned = re.sub(r"\s*:\s*$", "", cleaned)
    return cleaned


def normalize_visit_label(value: Any, fallback: Any = None) -> str | None:
    """Normalize visit values and sheet names to V1/V1A/V2/V3/V4... labels.

    ``V@`` is interpreted as ``V2`` because ``@`` is the shifted keyboard symbol
    for the number 2, a common spreadsheet naming typo.
    """
    raw = _plain_text(value)
    if not raw and fallback is not None:
        raw = _plain_text(fallback)
    if not raw:
        return None

    text = raw.lower().replace("@", "2")
    text = re.sub(r"[_-]+", " ", text)
    text = text.replace("baseline", "visit 1")
    compact = re.sub(r"[^a-z0-9]+", "", text)

    if any(token in compact for token in ("visit1and1a", "visit11a", "v1and1a", "v11a")):
        return "V1/V1A"
    if re.search(r"(?:visit|v)?1\s*(?:&|and|\+|/)\s*(?:visit|v)?1a", text):
        return "V1/V1A"
    if re.search(r"(?:visit|v)\s*1a\b", text) or compact in {"1a", "v1a", "visit1a"}:
        return "V1A"

    match = re.search(r"(?:visit|v)\s*0*([1-9]\d*)\b", text)
    if not match and compact.isdigit():
        match = re.match(r"0*([1-9]\d*)", compact)
    if match:
        return f"V{int(match.group(1))}"

    return None


def visit_sort_key(label: Any) -> tuple[float, str]:
    value = _plain_text(label).upper()
    if value == "V1/V1A":
        return (1.0, value)
    if value == "V1A":
        return (1.1, value)
    match = re.fullmatch(r"V(\d+)", value)
    if match:
        return (float(match.group(1)), value)
    return (999.0, value)


def ordered_visits(values: Iterable[Any]) -> list[str]:
    unique = {_plain_text(v) for v in values if _plain_text(v)}
    return sorted(unique, key=visit_sort_key)


def _normalize_missing_scalar(value: Any) -> Any:
    if value is None:
        return pd.NA
    if isinstance(value, str):
        cleaned = _plain_text(value)
        if cleaned.lower() in MISSING_TOKENS:
            return pd.NA
        return cleaned
    return value


def normalize_missing_values(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        if result[column].dtype == "object":
            result[column] = result[column].map(_normalize_missing_scalar)
    return result


def _normalise_patient_id(value: Any) -> Any:
    if value is None or pd.isna(value):
        return pd.NA
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    text = _plain_text(value)
    return text if text else pd.NA


def _values_conflict(row: pd.Series) -> bool:
    values = []
    for value in row:
        if pd.isna(value):
            continue
        if isinstance(value, pd.Timestamp):
            values.append(value.isoformat())
        else:
            values.append(_plain_text(value))
    return len(set(values)) > 1


def coalesce_and_canonicalize_columns(
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    """Canonicalize headers without discarding conflicting duplicate fields.

    Columns that normalize to the same canonical label are coalesced only when
    their non-missing values agree. If they contain conflicting values, each
    source column is retained with a numbered suffix so the dashboard never
    hides or overwrites source information.
    """
    groups: dict[str, list[int]] = {}
    originals: list[str] = []
    for index, original in enumerate(df.columns):
        canonical = canonical_column_name(original)
        if not canonical:
            canonical = f"Unnamed column {index + 1}"
        groups.setdefault(canonical, []).append(index)
        originals.append(_plain_text(original))

    output: dict[str, pd.Series] = {}
    mappings: list[dict[str, Any]] = []
    conflict_count = 0

    for canonical, indices in groups.items():
        block = pd.concat([df.iloc[:, idx] for idx in indices], axis=1)
        conflicts = int(block.apply(_values_conflict, axis=1).sum()) if len(indices) > 1 else 0
        conflict_count += conflicts

        if len(indices) == 1:
            final_names = [canonical]
            output[canonical] = block.iloc[:, 0]
            coalesced = False
        elif conflicts == 0:
            with pd.option_context("future.no_silent_downcasting", True):
                output[canonical] = block.bfill(axis=1).iloc[:, 0].infer_objects(copy=False)
            final_names = [canonical] * len(indices)
            coalesced = True
        else:
            # Preserve every conflicting source column. The first column keeps
            # the canonical name so standard metrics remain easy to select.
            final_names = []
            for position, idx in enumerate(indices, start=1):
                final_name = canonical if position == 1 else f"{canonical} [{position}]"
                output[final_name] = df.iloc[:, idx]
                final_names.append(final_name)
            coalesced = False

        for source_position, idx in enumerate(indices):
            final_name = final_names[source_position] if source_position < len(final_names) else canonical
            mappings.append(
                {
                    "Original Column": originals[idx],
                    "Canonical Column": final_name,
                    "Changed": originals[idx] != final_name,
                    "Coalesced": coalesced,
                    "Conflicting Rows": conflicts,
                }
            )

    mapping_df = pd.DataFrame(mappings)
    return pd.DataFrame(output), mapping_df, conflict_count


def convert_inferred_types(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        if column == "Patient ID":
            result[column] = result[column].map(_normalise_patient_id)
            continue
        if pd.api.types.is_datetime64_any_dtype(result[column]) or pd.api.types.is_numeric_dtype(result[column]):
            continue

        signature = _header_signature(column)
        non_null = result[column].dropna()
        if non_null.empty:
            continue

        if "date" in signature or signature.endswith(" time"):
            parsed_dates = pd.to_datetime(result[column], errors="coerce")
            if parsed_dates.notna().sum() / len(non_null) >= 0.55:
                result[column] = parsed_dates
                continue

        # Remove common display characters before testing numeric conversion.
        as_text = result[column].astype("string")
        cleaned = (
            as_text.str.replace(",", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.replace(r"^\s*<\s*", "", regex=True)
            .str.replace(r"^\s*>\s*", "", regex=True)
        )
        numeric = pd.to_numeric(cleaned, errors="coerce")
        if numeric.notna().sum() / len(non_null) >= 0.85:
            result[column] = numeric

    return result


def inspect_workbook(file_bytes: bytes) -> list[SheetInfo]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    infos: list[SheetInfo] = []
    try:
        for worksheet in workbook.worksheets:
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                tuple(),
            )
            infos.append(
                SheetInfo(
                    name=worksheet.title,
                    headers=tuple(header_row),
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                )
            )
    finally:
        workbook.close()
    return infos


def _last_nonempty_header_index(headers: tuple[Any, ...]) -> int:
    positions = [idx for idx, value in enumerate(headers) if _plain_text(value)]
    return positions[-1] if positions else -1


def _contains_id_and_event(headers: tuple[Any, ...]) -> bool:
    canonical = {canonical_column_name(value) for value in headers if _plain_text(value)}
    return "Patient ID" in canonical and "Event Name" in canonical


def _is_base_sheet(name: str) -> bool:
    compact = _plain_text(name).lower().replace("_", " ")
    return any(hint in compact for hint in BASE_SHEET_HINTS)


def _load_sheet(
    excel: pd.ExcelFile,
    info: SheetInfo,
    max_visit_columns: int = 2500,
) -> tuple[pd.DataFrame, str | None]:
    last_header = _last_nonempty_header_index(info.headers)
    if last_header < 0:
        return pd.DataFrame(), "Sheet has no header values."

    warning = None
    if last_header + 1 > max_visit_columns:
        # Protect the dashboard from malformed Excel dimensions or thousands of
        # generic placeholder columns. This path is not used for the supplied
        # visit sheets, but keeps future uploads responsive.
        meaningful = [
            idx
            for idx, value in enumerate(info.headers)
            if _plain_text(value)
            and not re.fullmatch(r"column\s*\d+", _plain_text(value), flags=re.I)
        ]
        if meaningful:
            usecols = sorted(set(meaningful))
            warning = (
                f"{info.name} had {last_header + 1:,} header columns; "
                f"loaded {len(usecols):,} non-generic columns to avoid an oversized range."
            )
        else:
            usecols = list(range(max_visit_columns))
            warning = (
                f"{info.name} had an oversized range; loading was capped at "
                f"{max_visit_columns:,} columns."
            )
    else:
        usecols = list(range(last_header + 1))

    frame = excel.parse(sheet_name=info.name, usecols=usecols, dtype=object)
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return frame, warning


def _find_visit_sheets(
    excel: pd.ExcelFile,
    infos: list[SheetInfo],
) -> tuple[list[SheetInfo], list[SheetInfo]]:
    visit_infos = [info for info in infos if normalize_visit_label(info.name) is not None]
    if visit_infos:
        excluded = [info for info in infos if info not in visit_infos]
        return visit_infos, excluded

    # Fallback for workbooks where sheets have generic names but Event Name
    # contains visit labels. Prefer non-base sheets, but support a single master
    # sheet when it is the only longitudinal source available.
    non_base_candidates: list[SheetInfo] = []
    base_candidates: list[SheetInfo] = []
    excluded: list[SheetInfo] = []
    for info in infos:
        if not _contains_id_and_event(info.headers):
            excluded.append(info)
            continue
        header_names = [canonical_column_name(value) for value in info.headers]
        event_index = header_names.index("Event Name")
        sample = excel.parse(
            sheet_name=info.name,
            usecols=[event_index],
            nrows=30,
            dtype=object,
        )
        normalized = sample.iloc[:, 0].map(normalize_visit_label)
        if normalized.notna().mean() >= 0.50:
            if _is_base_sheet(info.name):
                base_candidates.append(info)
            else:
                non_base_candidates.append(info)
        else:
            excluded.append(info)

    selected = non_base_candidates or base_candidates
    excluded.extend([info for info in base_candidates if info not in selected])
    return selected, excluded


def _date_candidate_indices(headers: tuple[Any, ...]) -> list[int]:
    candidates: list[tuple[int, int]] = []
    for idx, value in enumerate(headers):
        signature = _header_signature(value)
        if not signature:
            continue
        if "date" not in signature or "visit" not in signature:
            continue
        if any(term in signature for term in ("birth", "diagnos", "collection", "cgm", "start", "stop")):
            continue
        priority = 3
        if signature == "date of visit":
            priority = 0
        elif signature.startswith("date of the study visit"):
            priority = 1
        elif "study visit" in signature:
            priority = 2
        candidates.append((priority, idx))
    return [idx for _, idx in sorted(candidates)]


def _load_base_visit_date_lookup(
    excel: pd.ExcelFile,
    base_infos: list[SheetInfo],
) -> tuple[pd.DataFrame, str | None]:
    if not base_infos:
        return pd.DataFrame(columns=["Patient ID", "Visit", "Visit Date"]), None

    # Prefer explicitly named base/master sheets, then the sheet with most rows.
    ordered = sorted(
        base_infos,
        key=lambda info: (not _is_base_sheet(info.name), -info.max_row),
    )
    for info in ordered:
        canonical_headers = [canonical_column_name(value) for value in info.headers]
        if "Patient ID" not in canonical_headers or "Event Name" not in canonical_headers:
            continue
        date_indices = _date_candidate_indices(info.headers)
        if not date_indices:
            continue

        id_index = canonical_headers.index("Patient ID")
        event_index = canonical_headers.index("Event Name")
        usecols = sorted(set([id_index, event_index, *date_indices]))
        raw = excel.parse(sheet_name=info.name, usecols=usecols, dtype=object)
        raw = normalize_missing_values(raw)

        id_column = next(
            (column for column in raw.columns if canonical_column_name(column) == "Patient ID"),
            None,
        )
        event_column = next(
            (column for column in raw.columns if canonical_column_name(column) == "Event Name"),
            None,
        )
        date_columns = [
            column
            for column in raw.columns
            if canonical_column_name(column) == "Visit Date"
        ]
        if id_column is None or event_column is None or not date_columns:
            continue

        parsed_dates = pd.concat(
            [pd.to_datetime(raw[column], errors="coerce") for column in date_columns],
            axis=1,
        )
        # Visit-date fields are intentionally repeated by event in wide REDCap
        # exports. Select the first populated date in worksheet column order.
        with pd.option_context("future.no_silent_downcasting", True):
            visit_date = parsed_dates.bfill(axis=1).iloc[:, 0]

        lookup = pd.DataFrame(
            {
                "Patient ID": raw[id_column].map(_normalise_patient_id),
                "Visit": raw[event_column].map(normalize_visit_label),
                "Visit Date": visit_date,
            }
        ).dropna(subset=["Patient ID", "Visit"])
        lookup = lookup.sort_values(["Patient ID", "Visit"]).drop_duplicates(
            ["Patient ID", "Visit"], keep="first"
        )
        return lookup, info.name
    return pd.DataFrame(columns=["Patient ID", "Visit", "Visit Date"]), None


def _natural_id_key(value: Any) -> tuple[Any, ...]:
    text = _plain_text(value)
    parts = re.split(r"(\d+)", text)
    return tuple(int(part) if part.isdigit() else part.lower() for part in parts)


def load_clinical_workbook(file_bytes: bytes, file_name: str = "uploaded.xlsx") -> dict[str, Any]:
    """Load visit sheets and return a cleaned longitudinal data bundle."""
    infos = inspect_workbook(file_bytes)
    excel = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    visit_infos, excluded_infos = _find_visit_sheets(excel, infos)

    if not visit_infos:
        raise ValueError(
            "No visit sheets could be detected. Name sheets like V1, Visit_2, v3, "
            "or include an Event Name column containing visit labels."
        )

    frames: list[pd.DataFrame] = []
    sheet_rows: list[dict[str, Any]] = []
    mapping_frames: list[pd.DataFrame] = []
    warnings: list[str] = []
    total_conflicts = 0

    for info in visit_infos:
        raw, warning = _load_sheet(excel, info)
        if warning:
            warnings.append(warning)
        if raw.empty:
            warnings.append(f"{info.name} was detected as a visit sheet but contained no data rows.")
            continue

        raw_rows = len(raw)
        raw_columns = len(raw.columns)
        raw = normalize_missing_values(raw)
        cleaned, mapping, conflicts = coalesce_and_canonicalize_columns(raw)
        cleaned = convert_inferred_types(cleaned)
        total_conflicts += conflicts

        mapping.insert(0, "Sheet", info.name)
        mapping_frames.append(mapping)

        if "Patient ID" not in cleaned.columns:
            warnings.append(f"Skipped {info.name}: no patient ID column could be detected.")
            continue

        sheet_visit = normalize_visit_label(info.name)
        if "Event Name" in cleaned.columns:
            row_visits = cleaned["Event Name"].map(normalize_visit_label)
            cleaned["Visit"] = (
                row_visits.fillna(sheet_visit) if sheet_visit is not None else row_visits
            )
        else:
            cleaned["Event Name"] = pd.NA
            cleaned["Visit"] = sheet_visit

        cleaned["Source Sheet"] = info.name
        cleaned["Source Row"] = np.arange(2, len(cleaned) + 2)
        frames.append(cleaned)
        sheet_rows.append(
            {
                "Sheet": info.name,
                "Normalized Visit": sheet_visit,
                "Rows Loaded": raw_rows,
                "Columns Loaded": raw_columns,
                "Missing Patient IDs": int(cleaned["Patient ID"].isna().sum()),
            }
        )

    if not frames:
        raise ValueError("Visit sheets were found, but no usable patient rows could be loaded.")

    data = pd.concat(frames, ignore_index=True, sort=False)
    data["Patient ID"] = data["Patient ID"].map(_normalise_patient_id)

    base_date_lookup, base_date_sheet = _load_base_visit_date_lookup(excel, excluded_infos)
    if not base_date_lookup.empty:
        if "Visit Date" in data.columns:
            data = data.merge(
                base_date_lookup.rename(columns={"Visit Date": "Visit Date from Base"}),
                on=["Patient ID", "Visit"],
                how="left",
            )
            data["Visit Date"] = pd.to_datetime(data["Visit Date"], errors="coerce").fillna(
                pd.to_datetime(data["Visit Date from Base"], errors="coerce")
            )
            data = data.drop(columns=["Visit Date from Base"])
        else:
            data = data.merge(base_date_lookup, on=["Patient ID", "Visit"], how="left")
    elif "Visit Date" not in data.columns:
        data["Visit Date"] = pd.NaT

    data["Duplicate Patient-Visit"] = False
    valid_keys = data["Patient ID"].notna() & data["Visit"].notna()
    data.loc[valid_keys, "Duplicate Patient-Visit"] = data.loc[valid_keys].duplicated(
        ["Patient ID", "Visit"], keep=False
    )

    # Stable clinical ordering while keeping source-row traceability.
    patient_order = {
        patient_id: idx
        for idx, patient_id in enumerate(
            sorted(data["Patient ID"].dropna().unique(), key=_natural_id_key)
        )
    }
    data["_patient_order"] = data["Patient ID"].map(patient_order).fillna(10**9)
    data["_visit_order"] = data["Visit"].map(lambda value: visit_sort_key(value)[0])
    data = data.sort_values(
        ["_patient_order", "_visit_order", "Source Sheet", "Source Row"],
        kind="stable",
    ).drop(columns=["_patient_order", "_visit_order"])
    data = data.reset_index(drop=True)

    visit_counts = (
        data.groupby("Visit", dropna=False)
        .size()
        .rename("Rows")
        .reset_index()
    )
    visit_counts["Visit"] = visit_counts["Visit"].fillna("Unrecognized")
    visit_counts["_order"] = visit_counts["Visit"].map(lambda value: visit_sort_key(value)[0])
    visit_counts = visit_counts.sort_values(["_order", "Visit"]).drop(columns="_order")

    changed_mappings = pd.concat(mapping_frames, ignore_index=True) if mapping_frames else pd.DataFrame()
    if not changed_mappings.empty:
        changed_mappings = changed_mappings[
            changed_mappings["Changed"] | changed_mappings["Coalesced"]
        ].reset_index(drop=True)

    quality = {
        "File": file_name,
        "Visit Sheets Loaded": len(sheet_rows),
        "Excluded/Reference Sheets": len(excluded_infos),
        "Rows Loaded": int(len(data)),
        "Unique Patients": int(data["Patient ID"].nunique(dropna=True)),
        "Unique Patient-Visits": int(
            data.dropna(subset=["Patient ID", "Visit"])
            .drop_duplicates(["Patient ID", "Visit"])
            .shape[0]
        ),
        "Missing Patient IDs": int(data["Patient ID"].isna().sum()),
        "Duplicate Patient-Visit Rows": int(data["Duplicate Patient-Visit"].sum()),
        "Column Conflicts During Coalescing": int(total_conflicts),
        "Visit Date Coverage %": float(data["Visit Date"].notna().mean() * 100),
        "Base Date Source": base_date_sheet or "Not detected",
    }

    sheet_summary = pd.DataFrame(sheet_rows)
    excluded_summary = pd.DataFrame(
        [
            {
                "Sheet": info.name,
                "Rows": max(info.max_row - 1, 0),
                "Reported Columns": info.max_column,
                "Role": "Reference/base" if _contains_id_and_event(info.headers) else "Excluded",
            }
            for info in excluded_infos
        ]
    )

    result = {
        "data": data,
        "quality": quality,
        "sheet_summary": sheet_summary,
        "visit_counts": visit_counts,
        "column_mappings": changed_mappings,
        "excluded_sheets": excluded_summary,
        "warnings": warnings,
        "visit_order": ordered_visits(data["Visit"]),
    }
    excel.close()
    return result


def deduplicate_patient_visits(data: pd.DataFrame, method: str) -> pd.DataFrame:
    """Apply a user-selected duplicate policy without changing source data."""
    if method == "Keep all rows":
        return data.copy()

    keys = ["Patient ID", "Visit"]
    valid = data["Patient ID"].notna() & data["Visit"].notna()
    keyed = data.loc[valid].copy()
    unkeyed = data.loc[~valid].copy()

    if method == "Keep most complete row":
        value_columns = [column for column in keyed.columns if column not in META_COLUMNS]
        keyed["_completeness"] = keyed[value_columns].notna().sum(axis=1)
        keyed = (
            keyed.sort_values(keys + ["_completeness"], ascending=[True, True, False])
            .drop_duplicates(keys, keep="first")
            .drop(columns="_completeness")
        )
        return pd.concat([keyed, unkeyed], ignore_index=True, sort=False)

    if method == "Combine duplicate rows":
        def first_non_missing(series: pd.Series) -> Any:
            non_missing = series.dropna()
            return non_missing.iloc[0] if not non_missing.empty else pd.NA

        aggregations = {column: first_non_missing for column in keyed.columns if column not in keys}
        combined = keyed.groupby(keys, as_index=False, dropna=False).agg(aggregations)
        combined["Duplicate Patient-Visit"] = False
        return pd.concat([combined, unkeyed], ignore_index=True, sort=False)

    return data.copy()


def numeric_metric_columns(data: pd.DataFrame, minimum_values: int = 3) -> list[str]:
    metrics = []
    for column in data.columns:
        if column in META_COLUMNS:
            continue
        if pd.api.types.is_numeric_dtype(data[column]) and data[column].notna().sum() >= minimum_values:
            metrics.append(column)
    return metrics


def suggested_metrics(data: pd.DataFrame, limit: int = 6) -> list[str]:
    numeric = numeric_metric_columns(data)
    selected = [metric for metric in DEFAULT_METRIC_PRIORITY if metric in numeric]
    if len(selected) < limit:
        selected.extend([metric for metric in numeric if metric not in selected][: limit - len(selected)])
    return selected[:limit]


def completion_rate(data: pd.DataFrame, visits: list[str] | None = None) -> float:
    valid = data.dropna(subset=["Patient ID", "Visit"])
    if visits:
        valid = valid[valid["Visit"].isin(visits)]
    patient_count = valid["Patient ID"].nunique()
    visit_count = len(visits or ordered_visits(valid["Visit"]))
    if patient_count == 0 or visit_count == 0:
        return 0.0
    observed = valid.drop_duplicates(["Patient ID", "Visit"]).shape[0]
    return observed / (patient_count * visit_count) * 100


def summary_statistics(data: pd.DataFrame, metrics: list[str]) -> pd.DataFrame:
    available = [metric for metric in metrics if metric in data.columns]
    if not available or data.empty:
        return pd.DataFrame(columns=["Visit", "Metric", "N", "Mean", "Median", "SD"])

    rows = []
    for visit in ordered_visits(data["Visit"]):
        visit_data = data[data["Visit"] == visit]
        for metric in available:
            values = pd.to_numeric(visit_data[metric], errors="coerce").dropna()
            rows.append(
                {
                    "Visit": visit,
                    "Metric": metric,
                    "N": int(values.count()),
                    "Mean": values.mean() if not values.empty else np.nan,
                    "Median": values.median() if not values.empty else np.nan,
                    "SD": values.std(ddof=1) if len(values) > 1 else np.nan,
                }
            )
    return pd.DataFrame(rows)


def missingness_table(data: pd.DataFrame, columns: list[str] | None = None) -> pd.DataFrame:
    selected = [column for column in (columns or list(data.columns)) if column in data.columns]
    if not selected:
        return pd.DataFrame(columns=["Column", "Missing", "Missing %", "Non-missing"])
    result = pd.DataFrame(
        {
            "Column": selected,
            "Missing": [int(data[column].isna().sum()) for column in selected],
            "Missing %": [float(data[column].isna().mean() * 100) for column in selected],
            "Non-missing": [int(data[column].notna().sum()) for column in selected],
        }
    )
    return result.sort_values(["Missing %", "Column"], ascending=[False, True]).reset_index(drop=True)


def filter_clinical_data(
    data: pd.DataFrame,
    patient_ids: list[str] | None = None,
    visits: list[str] | None = None,
    date_range: tuple[Any, Any] | None = None,
    numeric_ranges: dict[str, tuple[float, float]] | None = None,
) -> pd.DataFrame:
    filtered = data.copy()
    if patient_ids:
        filtered = filtered[filtered["Patient ID"].isin(patient_ids)]
    if visits:
        filtered = filtered[filtered["Visit"].isin(visits)]
    if date_range and "Visit Date" in filtered.columns:
        start, end = pd.Timestamp(date_range[0]), pd.Timestamp(date_range[1])
        dates = pd.to_datetime(filtered["Visit Date"], errors="coerce")
        filtered = filtered[dates.between(start, end, inclusive="both")]
    for column, (lower, upper) in (numeric_ranges or {}).items():
        if column not in filtered.columns:
            continue
        values = pd.to_numeric(filtered[column], errors="coerce")
        filtered = filtered[values.between(lower, upper, inclusive="both") | values.isna()]
    return filtered.reset_index(drop=True)


def dataframe_to_excel_bytes(
    data: pd.DataFrame,
    stats: pd.DataFrame | None = None,
    quality: dict[str, Any] | None = None,
) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="Filtered Data", index=False)
        if stats is not None and not stats.empty:
            stats.to_excel(writer, sheet_name="Summary Statistics", index=False)
        if quality:
            pd.DataFrame(
                [{"Check": key, "Value": value} for key, value in quality.items()]
            ).to_excel(writer, sheet_name="Data Quality", index=False)
    return buffer.getvalue()
